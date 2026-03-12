# -*- coding: latin-1 -*-
#! / usr / bin / python
# vim: set fileencoding = latin-1:
import os, sys
import subprocess
import unicodedata
import tkinter.filedialog  #Libreria para crear cuadros de dialogo
import csv, sys, re, datetime, os #Librerias de python y del sistema
import xlwt, xlrd     #Librerias de gestion de archivos excel
import sys
import openpyxl
import math
import pandas as pd
import datetime
from openpyxl.utils import get_column_letter
import re
from dateutil import parser
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'Script_validacion_caracteristicas_de_paquetes')))
from ConsultaGeneralPaquetes import Paquetes   #impartando funciones de modulo de BD
from ConsultaBonosPorParent import BonosPorParent #impartando funciones de modulo de BD
from ConsultaPaquetesBono import BonosPorPaqueteDetallado #impartando funciones de modulo de BD
from openpyxl.workbook import Workbook as openpyxlWorkbook
from openpyxl.styles import PatternFill, Alignment, Font

lista_id_paquetes = [] #lista de equipos de la base
lista_id_bonos_paquetes = []
lista_planilla = [] #planilla de MKT con caracteristicas especificadas 
lista_rangos = []
lista_loyalty = []
tipo = ""
#estilo del Excel de salida
workbook2 = xlwt.Workbook('ascii',style_compression=2) 
xlwt.add_palette_colour("purpura_custom", 0x21)
workbook2.set_colour_RGB(0x21, 204, 204, 255)
xlwt.add_palette_colour("verde_custom", 0x22)
workbook2.set_colour_RGB(0x22, 0, 255, 0)
xlwt.add_palette_colour("red_custom", 0x23)
workbook2.set_colour_RGB(0x23, 255, 102, 0)
xlwt.add_palette_colour("azul_custom", 0x24)
workbook2.set_colour_RGB(0x24, 0, 204, 255)
xlwt.add_palette_colour("amarillo_custom", 0x25)
workbook2.set_colour_RGB(0x25, 255, 255, 153)
sheet_bonos = workbook2.add_sheet("Bonos",  cell_overwrite_ok=True)
sheet_paquetes = workbook2.add_sheet("Paquetes",  cell_overwrite_ok=True)

def buscarPosiciones(titulo):
    global lista_planilla
    for i in range(len(lista_planilla[3])):
        if str(titulo).lower() == str(lista_planilla[3][i]).lower():
            #print(titulo,i)
            return i   
    mensaje = "No se encontro el titulo " + titulo + " en la BDD"
    sys.exit(mensaje)

def openFileDialog():
    """Esta funcion abre un cuadro de dialogo para cargar archivos."""
    root = tkinter.Tk()
    root.withdraw()
    file_path = tkinter.filedialog.askopenfilename()
    if not file_path:  # Verificar si el usuario no seleccionó ningún archivo
        print("No se selecciono ningun archivo.")
        exit()  # Terminar el programa si no se seleccionó ningún archivo
    return file_path
    
def saveFileDialog():
    '''Esta funcion abre un cuadro de dialogo para guardar archivos''' 
    ftypes = [ ('Libro de Excel','*.xlsx'), ('Libro de Excel 97-2003','*.xls'), ('All files', '*'),  ]
    root = tkinter.Tk()
    root.withdraw()
    file_path = tkinter.filedialog.asksaveasfilename(filetypes=ftypes)
    return file_path + ".xls"

def readCSVPlanes(file_path):
    # Crear una lista vacía para almacenar las filas como listas
    lista_filas = []
    # Cargar el libro de trabajo de Excel
    wb = openpyxl.load_workbook(file_path, data_only=True)
    # Obtener la hoja activa
    sheet = wb.active
    # Iterar sobre las filas de la hoja
    for row in sheet.iter_rows(values_only=True):
        # Agregar la fila como lista a la lista de filas
        lista_filas.append(list(row))
    # Cerrar el libro de trabajo para liberar recursos
    wb.close()
    # Devolver el DataFrame y la lista de filas
    return lista_filas

def guardarComoXLSX(filepath):
    """Guardar el archivo como .XLSX, borrar el .XLS generado y aplicar estilos/anchos/colores.

    Nota: el .XLS (xlwt) no preserva estilos al pasarlo a .XLSX, por eso ac� re-aplicamos:
    - Encabezado resaltado
    - Colores verde/rojo seg�n Result (OK/OKA/ERROR/etc.)
    - Wrap en columna de Error
    - Ajuste de anchos (con reglas especiales para Error)
    """
    xlsBook = xlrd.open_workbook(filepath)
    workbook2 = openpyxlWorkbook()

    # fills
    FILL_HEADER = PatternFill(fill_type="solid", start_color="00CCFF", end_color="00CCFF")  # celeste
    FILL_GREEN  = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")  # verde claro
    FILL_RED    = PatternFill(fill_type="solid", start_color="FFC7CE", end_color="FFC7CE")  # rojo claro

    for i in range(0, xlsBook.nsheets):
        xlsSheet = xlsBook.sheet_by_index(i)
        sheet2 = workbook2.active if i == 0 else workbook2.create_sheet()
        sheet2.title = xlsSheet.name

        # Copiar valores
        for row in range(0, xlsSheet.nrows):
            for col in range(0, xlsSheet.ncols):
                cell = sheet2.cell(row=row + 1, column=col + 1)
                cell.value = xlsSheet.cell_value(row, col)

        max_row = sheet2.max_row
        max_col = sheet2.max_column

        # Detectar columnas por encabezado (m�s robusto que hardcodear �ndices)
        headers = {}
        for c in range(1, max_col + 1):
            v = sheet2.cell(row=1, column=c).value
            key = str(v).strip().lower() if v is not None else ""
            headers[key] = c

        # posibles nombres de columnas
        def _find_col(posibles, default):
            for k in posibles:
                if k in headers:
                    return headers[k]
            # fallback: buscar por "contiene"
            for key, c in headers.items():
                for k in posibles:
                    if k and k in key:
                        return c
            return default

        col_result = _find_col(["result", "resultado", "estado"], 6)
        col_error  = _find_col(["error", "detalle error", "descripcion error", "descripci�n error", "descripcion"], 7)

        # Estilo encabezado (fila 1)
        for c in range(1, max_col + 1):
            h = sheet2.cell(row=1, column=c)
            h.font = Font(name='Calibri', size=11, color="000000", bold=True)
            h.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            h.fill = FILL_HEADER

        # Coloreo seg�n Result + wrap/altura en Error
        for r in range(2, max_row + 1):
            result_cell = sheet2.cell(row=r, column=col_result)
            error_cell  = sheet2.cell(row=r, column=col_error)

            # wrap siempre en la columna de error
            error_cell.alignment = Alignment(wrap_text=True, vertical="top")

            val = ""
            if result_cell.value is not None:
                val = str(result_cell.value).strip().lower()

            # aceptar OK, OKA, "ok - ..." y variantes
            is_ok = val.startswith("ok") or " oka" in f" {val} " or val == "oka"
            is_error = ("error" in val) or val.startswith("err")

            if is_ok:
                result_cell.fill = FILL_GREEN
                error_cell.fill = FILL_GREEN  # ayuda visual
            elif is_error:
                result_cell.fill = FILL_RED
                error_cell.fill = FILL_RED

            # ajustar altura de fila si hay mucho texto en error
            if error_cell.value is not None:
                t = str(error_cell.value)
                if len(t) > 60:
                    # estimaci�n: cu�ntas "l�neas" va a ocupar con un ancho razonable
                    est_lines = int(math.ceil(len(t) / 60.0))
                    sheet2.row_dimensions[r].height = min(160, 15 * est_lines)

        # Ajuste de ancho de columnas
        for c in range(1, max_col + 1):
            # Para la columna de error damos m�s ancho por defecto
            if c == col_error:
                sheet2.column_dimensions[get_column_letter(c)].width = 80
                continue

            max_len = 0
            for r in range(1, max_row + 1):
                v = sheet2.cell(row=r, column=c).value
                if v is None:
                    continue
                l = len(str(v))
                if l > max_len:
                    max_len = l

            sheet2.column_dimensions[get_column_letter(c)].width = min(60, max_len + 2)

        # Congelar encabezado
        sheet2.freeze_panes = "A2"

    filepath2 = filepath[:-4] + '.xlsx'
    workbook2.save(filepath2)

    if os.path.exists(filepath):
        os.remove(filepath)
    else:
        print("The file does not exist")

def ImprimirEncabezadoBonos(sheet):
    nombre_columnas = ['Offering Name', 'Bono', 'Codigo Bono', 'Amount', 'Descripcion', 'Result', 'Error']
    Font = "font: name Calibri, color-index black, height 220"
    for i in range(len(nombre_columnas)):
        sheet.write(0, i, nombre_columnas[i],
                    xlwt.easyxf("pattern: pattern solid, fore_color azul_custom; align: horiz center; " + Font))
    return 0

def ImprimirEncabezadoPaquetes(sheet):
    nombre_columnas = ['Paquete', 'Client Friendly Name', 'Modalidad de cobro',
                       'Action Case Code', 'Marketing Group', 'Result', 'Error']
    Font = "font: name Calibri, color-index black, height 220"
    for i in range(len(nombre_columnas)):
        sheet.write(0, i, nombre_columnas[i],
                    xlwt.easyxf("pattern: pattern solid, fore_color azul_custom; align: horiz center; " + Font))
    return 0

def ImprimirBaseBonos(sheet, row, offering_name, bono_nombre, codigo_bono, amount, descripcion):
    sheet.write(row, 0, str(offering_name))
    sheet.write(row, 1, str(bono_nombre))
    sheet.write(row, 2, str(codigo_bono))
    sheet.write(row, 3, str(amount))
    sheet.write(row, 4, str(descripcion))

def ImprimirBasePaquetes(sheet, row, offering_name, client_name, modalidad_de_cobro, action_case_code, marketing_group):
    sheet.write(row, 0, str(offering_name))
    sheet.write(row, 1, str(client_name))
    sheet.write(row, 2, str(modalidad_de_cobro))
    sheet.write(row, 3, "" if action_case_code is None else str(action_case_code))
    sheet.write(row, 4, str(marketing_group))
def extraer_valor(campo):
    """
    Recibe un string tipo 'Nombre: valor' y devuelve solo 'valor'
    sin espacios al inicio/fin y sin comillas.
    Si no hay valor (ej: 'Nombre:'), devuelve ''.
    """
    if campo is None:
        return ""

    # Asegurar que es string
    texto = str(campo)

    # Partir por el primer ':'
    partes = texto.split(":", 1)

    if len(partes) < 2:
        # No tiene ':', devolver string limpio
        return texto.strip()

    # Tomar lo que está después de los ':'
    valor = partes[1].strip()

    # Remover comillas simples si las hubiera
    if valor.startswith("'") and valor.endswith("'"):
        valor = valor[1:-1].strip()

    return valor
                                     
    
    
def ValidarDatosPaquetes(paquete_nombre,client_name,modalidad_de_cobro,marketing_group,action_cause_code,paquete):
    paquetes_detalles = _split_comas(paquete[5])
    # print(paquetes_detalles)
    sistema_nombre = paquete[2]
    sistmea_client = extraer_valor(paquetes_detalles[3])
    #print(sistmea_client)
    sistema_cobro = extraer_valor(paquetes_detalles[0])
    sistema_cobro = int(sistema_cobro)
    #print(sistema_cobro)
    sistema_group = extraer_valor(paquetes_detalles[9])
    #print(sistema_group)
    sistema_action_cause_code = extraer_valor(paquetes_detalles[1])
    print(sistema_action_cause_code)
    respuesta = ['OK', '']
    if(str(sistema_nombre).lower() != paquete_nombre):
        respuesta[0] = "Error"
        if(respuesta[1]):
            respuesta[1] += ", Distinto nombre ->"
            respuesta[1]+= str(paquete_nombre)
        else: 
            respuesta[1] = "Distinto nombre->"
            respuesta[1]+= str(paquete_nombre)   
    if(str(sistmea_client).lower() != client_name):
        respuesta[0] = "Error"
        if(respuesta[1]):
            respuesta[1] += ", Distinto client nombre ->"
            respuesta[1]+= str(client_name)
        else: 
            respuesta[1] = "Distinto client nombre->"
            respuesta[1]+= str(client_name)   
    if(sistema_cobro != modalidad_de_cobro):
        respuesta[0] = "Error"
        if(respuesta[1]):
            respuesta[1] += ", Distinta modalidad de cobro ->"
            respuesta[1]+= str(modalidad_de_cobro)
        else: 
            respuesta[1] = "Distinta modalidad de cobro->"
            respuesta[1]+= str(modalidad_de_cobro)   
    if(str(sistema_group).lower() != marketing_group):
        respuesta[0] = "Error"
        if(respuesta[1]):
            respuesta[1] += ", Distinto marketing group ->"
            respuesta[1]+= str(marketing_group)
        else: 
            respuesta[1] = "Distinto marketing group->"
            respuesta[1]+= str(marketing_group)   
    if(str(sistema_action_cause_code) != action_cause_code):
        respuesta[0] = "Error"
        if(respuesta[1]):
            respuesta[1] += ", Distinto action_cause_code ->"
            respuesta[1]+= str(action_cause_code)
        else: 
            respuesta[1] = "Distinto nombre->"
            respuesta[1]+= str(action_cause_code)                                                    

    return respuesta

def buscarBono(bono_nombre, paquete_nombre):
       # Normalizar entradas a minúsculas
    bono_nombre = str(bono_nombre) 
    paquete_nombre = str(paquete_nombre)
    #print(bono_nombre) 
    #print(paquete_nombre)
    for i in range(len(lista_id_bonos_paquetes)): 
        # Normalizar valorpaquetes a minúsculas para la comparación
        Pnombre = str(lista_id_bonos_paquetes[i][1])
        Bnombre = str(lista_id_bonos_paquetes[i][3])
        # Comparar con normalización y verificar si está en desarrollo
        if (bono_nombre == Bnombre and paquete_nombre == Pnombre):
            return lista_id_bonos_paquetes[i]
    
    # Formatear mensaje de error como una cadena
    #mensaje = 'error - No se encontro el Paquete'
    return None

def buscarPaquete(paquete_nombre):
       # Normalizar entradas a minúsculas
    paquete_nombre = str(paquete_nombre).lower() 
    for i in range(len(lista_id_paquetes)): 
        # Normalizar valorpaquetes a minúsculas para la comparación
        Pnombre = str(lista_id_paquetes[i][2]).lower()
        # Comparar con normalización y verificar si está en desarrollo
        if (paquete_nombre == Pnombre):
            return lista_id_paquetes[i]
    
    # Formatear mensaje de error como una cadena
    mensaje = 'error - No se encontro el Paquete'
    return None
            
def _split_comas(value):
    # Acepta str o lista/tupla; devuelve lista limpia de tokens
    if value is None:
        return []
    if isinstance(value, (list, tuple, set)):
        items = list(value)
    else:
        items = str(value).split(",")
    out = []
    for x in items:
        sx = str(x).strip()
        if sx:
            out.append(sx)
    return out           
            
            
def ValidarDatosBonos(bono_nombre, cantidad_bono,unidad_bono,id_bono,vigencia_bono,bono):
    sistema_nombre = bono[3]
    sistema_amount = bono[5]
    sistema_vigencia = bono[4]
    sistema_id_bono = bono[6]
    sistema_descripcion_bono = bono[7]
    sistema_min = bono[8]
    sistema_max = bono[9]
    sistema_default_behavior =bono[10]
    sistema_object_type = bono[12]
    cantidad_bono = float(cantidad_bono)
    respuesta = ['OK', '']
    if(str(sistema_nombre) != bono_nombre):
        respuesta[0] = "Error"
        if(respuesta[1]):
            respuesta[1] += ", Distinto nombre ->"
            respuesta[1]+= str(bono_nombre)
        else: 
            respuesta[1] = "Distinto nombre->"
            respuesta[1]+= str(bono_nombre) 
    if(int(sistema_vigencia) != int(vigencia_bono + 1)):
        respuesta[0] = "Error"
        if(respuesta[1]):
            respuesta[1] += ", Distinta vigencia ->"
            respuesta[1]+= str(vigencia_bono)
        else: 
            respuesta[1] = "Distinto nombre->"
            respuesta[1]+= str(vigencia_bono)
    if(str(sistema_id_bono).lower() != id_bono):
        respuesta[0] = "Error"
        if(respuesta[1]):
            respuesta[1] += ", Distinto id bono ->"
            respuesta[1]+= str(id_bono)
        else: 
            respuesta[1] = "Distinto id bono->"
            respuesta[1]+= str(id_bono) 
    if(str(sistema_descripcion_bono).lower() not in str(bono_nombre).lower()):
        respuesta[0] = "Error"
        if(respuesta[1]):
            respuesta[1] += ", Distinto bonus description ->"
            respuesta[1]+= str(bono_nombre)
        else: 
            respuesta[1] = "Distinto bonus description->"
            respuesta[1]+= str(bono_nombre) 
    if(unidad_bono == 'GB' and sistema_amount != None):
        if(float(str(sistema_amount)) != ((((cantidad_bono *1024)*1024)*1024))):
            print(float(str(sistema_amount)))
            print(((((cantidad_bono *1024)*1024)*1024)))
            respuesta[0] = "Error"
            if(respuesta[1]):
                respuesta[1] += ", Distinto amount ->"
                respuesta[1]+= str(cantidad_bono)
            else: 
                respuesta[1] = "Distinto amount->"
                respuesta[1]+= str(cantidad_bono)
    if(unidad_bono == 'MB' and sistema_amount != None):
        if(float(str(sistema_amount)) != (((cantidad_bono *1024)*1024))):
            respuesta[0] = "Error"
            if(respuesta[1]):
                respuesta[1] += ", Distinto amount ->"
                respuesta[1]+= str(cantidad_bono)
            else: 
                respuesta[1] = "Distinto amount->"
                respuesta[1]+= str(cantidad_bono)
    if(unidad_bono == 'TB' and sistema_amount != None):
        if(float(str(sistema_amount)) != ((((cantidad_bono *1024)*1024)*1024)*1024)):
            respuesta[0] = "Error"
            if(respuesta[1]):
                respuesta[1] += ", Distinto amount ->"
                respuesta[1]+= str(cantidad_bono)
            else: 
                respuesta[1] = "Distinto amount->"
                respuesta[1]+= str(cantidad_bono)                                 
    else:
        if((unidad_bono == 'minutos' or unidad_bono == 'Min'or unidad_bono == 'min') and sistema_amount != None):
            cantidad = (int(cantidad_bono) *60)
            if(int(str(sistema_amount)) != cantidad):
                # print(int(str(sistema_amount)))
                # print(cantidad)
                respuesta[0] = "Error"
                if(respuesta[1]):
                    respuesta[1] += ", Distinto amount ->"
                    respuesta[1]+= str(cantidad_bono)
                else: 
                    respuesta[1] = "Distinto amount->"
                    respuesta[1]+= str(cantidad_bono) 
        else:
            if (unidad_bono == 'sms' and sistema_amount != None):
                if(float(sistema_amount) != float(cantidad_bono)):
                    respuesta[0] = "Error"
                if(respuesta[1]):
                    respuesta[1] += ", Distinto amount ->"
                    respuesta[1]+= str(cantidad_bono)
                else: 
                    respuesta[1] = "Distinto amount->"
                    respuesta[1]+= str(cantidad_bono)
    if(int(sistema_max) != 1 or int(sistema_min) != 1 or sistema_default_behavior != 'Selected'):
        respuesta[0] = "Error"
        if(respuesta[1]):
            respuesta[1] += ", error en la relacion padre hijo"
        else: 
            respuesta[1] = "Error en la relacion padre hijo"
    if(sistema_object_type != 'Bolt-On'):
            respuesta[0] = "Error"
            if(respuesta[1]):
                respuesta[1] += ", Distinto object type ->"
                respuesta[1]+= str(sistema_object_type)
            else: 
                respuesta[1] = "Distinto object type->"
                respuesta[1]+= str(sistema_object_type)                         
                                                                                                 
    return respuesta   

def imprimirEquipos(CantidadBonos):
    '''Esta funcion crea el archivo excel con dos hojas: Bonos y Paquetes'''
    global workbook2, sheet_bonos, sheet_paquetes

    CantidadBonos = int(CantidadBonos)

    # =========================
    # Hoja BONOS (primera)
    # =========================
    ImprimirEncabezadoBonos(sheet_bonos)
    jb = 1  # fila de datos (0 es encabezado)

    for i in range(len(lista_planilla) - 3):
        i = i + 3
        k = 0

        for _ in range(CantidadBonos):
            bono_nombre_cell = lista_planilla[i][19 + (k * 5)]
            if bono_nombre_cell is None or str(bono_nombre_cell).strip() == "":
                k += 1
                continue

            bono_nombre = bono_nombre_cell
            cantidad_bono = lista_planilla[i][15 + (k * 5)]
            unidad_bono = lista_planilla[i][16 + (k * 5)]
            id_bono = lista_planilla[i][17 + (k * 5)]
            paquete_nombre = lista_planilla[i][4]
            vigencia_bono = lista_planilla[i][10]

            bono = buscarBono(bono_nombre, paquete_nombre)
            if bono is None:
                print('No se encontro el bono EXACTO:')
                print('  Excel paquete =', repr(paquete_nombre))
                print('  Excel bono    =', repr(bono_nombre))
                print('  Bonos en CR para ese paquete:')
                for fila in lista_id_bonos_paquetes:
                    if fila[1] == paquete_nombre:
                        print('    -', repr(fila[3]))
            else:
                # Mapeo de columnas en hoja Bonos:
                # Offering Name | Bono | Codigo Bono | Amount | Descripcion
                ImprimirBaseBonos(sheet_bonos, jb, bono[1], bono_nombre, id_bono, cantidad_bono, unidad_bono)

                Resultado = ValidarDatosBonos(str(bono_nombre), float(cantidad_bono), str(unidad_bono),
                                              str(id_bono).lower(), vigencia_bono, bono)
                sheet_bonos.write(jb, 5, Resultado[0])  # Result
                sheet_bonos.write(jb, 6, Resultado[1])  # Error
                jb += 1

            k += 1

    # =========================
    # Hoja PAQUETES (segunda)
    # =========================
    ImprimirEncabezadoPaquetes(sheet_paquetes)
    jp = 1  # fila de datos

    for p in range(len(lista_planilla) - 3):
        p = p + 3

        paquete_nombre = lista_planilla[p][4]
        client_name = lista_planilla[p][5]
        modalidad_de_cobro = lista_planilla[p][11]

        # Seg�n tu planilla: el marketing group depende de la cantidad de bonos
        marketing_group_idx = ((CantidadBonos + 1) * 5) + 14
        marketing_group = lista_planilla[p][marketing_group_idx] if marketing_group_idx < len(lista_planilla[p]) else ""
        action_cause_code = lista_planilla[p][((CantidadBonos + 1) * 5) + 22]
        paquete = buscarPaquete(paquete_nombre)
        if paquete is None:
            print('No se encontro el paquete', paquete_nombre)
        else:
            ImprimirBasePaquetes(sheet_paquetes, jp, paquete_nombre, client_name, modalidad_de_cobro, action_cause_code, marketing_group)

            Resultado = ValidarDatosPaquetes(str(paquete_nombre).lower(), client_name, int(modalidad_de_cobro),
                                             str(marketing_group).lower(), str(action_cause_code), paquete)
            sheet_paquetes.write(jp, 5, Resultado[0])  # Result
            sheet_paquetes.write(jp, 6, Resultado[1])  # Error
            jp += 1

    ruta = saveFileDialog()
    workbook2.save(ruta)
    guardarComoXLSX(ruta)
def guardarArchivoText(nombre, objeto):
    '''Funcion guarda archivo en formato texto'''
    archivo = open(nombre,'w') 
    str1 = ',\n'.join(str(e) for e in objeto)
    archivo.write(str1)  
    archivo.close() 
    
    
    
if __name__ == '__main__':
    """Esta la funcion principal""" 
        # Seleccionar opcion para ejecutar script 
    print ('Por favor introduzca el ID del CR:')
    CRID = input()
    print('Introducir cantidad de bonos')
    cantidad_bonos = input()
    lista_id_paquetes = Paquetes(CRID)
    lista_id_bonos = BonosPorParent(CRID)
    lista_id_bonos_paquetes = BonosPorPaqueteDetallado(CRID)
    #lista_id_loyalty = Loyalty(estado)
    now = datetime.datetime.now()
    fecha = now.strftime("%Y-%m-%d %H.%M")
    archivo3 = 'Lista de Paquetes en CR'+fecha+'.txt'
    #guardarArchivoText(archivo2, lista_id_equipos_Online)
    guardarArchivoText(archivo3, lista_id_paquetes)
    archivo2 = 'Lista de Bonos en CR'+fecha+'.txt'
    #guardarArchivoText(archivo2, lista_id_equipos_Online)
    guardarArchivoText(archivo2, lista_id_bonos)
    archivo1 = 'Lista de BonosPaquetes en CR'+fecha+'.txt'
    #guardarArchivoText(archivo2, lista_id_equipos_Online)
    guardarArchivoText(archivo1, lista_id_bonos_paquetes)
    #Solicitar excel de nuevos precios
    ruta = openFileDialog()
    lista_planilla = readCSVPlanes(ruta)    
    imprimirEquipos(cantidad_bonos)    
    print ('Fin')    
       