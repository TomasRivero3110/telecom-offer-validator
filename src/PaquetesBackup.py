# -*- coding: latin-1 -*-
#! / usr / bin / python
# vim: set fileencoding = latin-1:
import os, sys
import subprocess
import unicodedata
import tkinter.filedialog  #Libreria para crear cuadros de dialogo
import csv, sys, re, datetime, os #Librerias de python y del sistema
import xlwt, xlrd     #Librerias de gestion de archivos excel
import sys
import openpyxl
import pandas as pd
import datetime
from openpyxl.utils import get_column_letter
import re
from dateutil import parser
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'Script_validacion_caracteristicas_de_paquetes')))
from ConsultaGeneralPaquetes import Paquetes   #impartando funciones de modulo de BD
from ConsultaBonosPorParent import BonosPorParent #impartando funciones de modulo de BD
from ConsultaPaquetesBono import BonosPorPaqueteDetallado #impartando funciones de modulo de BD
from openpyxl.workbook import Workbook as openpyxlWorkbook
from openpyxl.styles import PatternFill, Alignment, Font

lista_id_paquetes = [] #lista de equipos de la base
lista_id_bonos_paquetes = []
lista_planilla = [] #planilla de MKT con caracteristicas especificadas 
lista_rangos = []
lista_loyalty = []
tipo = ""
#estilo del Excel de salida
workbook2 = xlwt.Workbook('ascii',style_compression=2) 
xlwt.add_palette_colour("purpura_custom", 0x21)
workbook2.set_colour_RGB(0x21, 204, 204, 255)
xlwt.add_palette_colour("verde_custom", 0x22)
workbook2.set_colour_RGB(0x22, 0, 255, 0)
xlwt.add_palette_colour("red_custom", 0x23)
workbook2.set_colour_RGB(0x23, 255, 102, 0)
xlwt.add_palette_colour("azul_custom", 0x24)
workbook2.set_colour_RGB(0x24, 0, 204, 255)
xlwt.add_palette_colour("amarillo_custom", 0x25)
workbook2.set_colour_RGB(0x25, 255, 255, 153)
sheet2 = workbook2.add_sheet("POI Worksheet",  cell_overwrite_ok=True)

def buscarPosiciones(titulo):
    global lista_planilla
    for i in range(len(lista_planilla[3])):
        if str(titulo).lower() == str(lista_planilla[3][i]).lower():
            #print(titulo,i)
            return i   
    mensaje = "No se encontro el titulo " + titulo + " en la BDD"
    sys.exit(mensaje)

def openFileDialog():
    """Esta funcion abre un cuadro de dialogo para cargar archivos."""
    root = tkinter.Tk()
    root.withdraw()
    file_path = tkinter.filedialog.askopenfilename()
    if not file_path:  # Verificar si el usuario no seleccionó ningún archivo
        print("No se selecciono ningun archivo.")
        exit()  # Terminar el programa si no se seleccionó ningún archivo
    return file_path
    
def saveFileDialog():
    '''Esta funcion abre un cuadro de dialogo para guardar archivos''' 
    ftypes = [ ('Libro de Excel','*.xlsx'), ('Libro de Excel 97-2003','*.xls'), ('All files', '*'),  ]
    root = tkinter.Tk()
    root.withdraw()
    file_path = tkinter.filedialog.asksaveasfilename(filetypes=ftypes)
    return file_path + ".xls"

def readCSVPlanes(file_path):
    # Crear una lista vacía para almacenar las filas como listas
    lista_filas = []
    # Cargar el libro de trabajo de Excel
    wb = openpyxl.load_workbook(file_path, data_only=True)
    # Obtener la hoja activa
    sheet = wb.active
    # Iterar sobre las filas de la hoja
    for row in sheet.iter_rows(values_only=True):
        # Agregar la fila como lista a la lista de filas
        lista_filas.append(list(row))
    # Cerrar el libro de trabajo para liberar recursos
    wb.close()
    # Devolver el DataFrame y la lista de filas
    return lista_filas

def guardarComoXLSX(filepath):
    '''Guardar el archivo como .XLSX, borrar el .XLS generado y aplicar estilos/anchos'''
    xlsBook = xlrd.open_workbook(filepath)
    workbook2 = openpyxlWorkbook()
    sheet2 = None
    xrange = range

    for i in xrange(0, xlsBook.nsheets):
        xlsSheet = xlsBook.sheet_by_index(i)
        sheet2 = workbook2.active if i == 0 else workbook2.create_sheet()
        sheet2.title = xlsSheet.name

        # Copiar valores
        for row in xrange(0, xlsSheet.nrows):
            for col in xrange(0, xlsSheet.ncols):
                cell = sheet2.cell(row=row + 1, column=col + 1)
                cell.value = xlsSheet.cell_value(row, col)

        # Estilos de encabezado (fila 1)
        max_row = sheet2.max_row
        max_col = sheet2.max_column
        for c in range(1, max_col + 1):
            h = sheet2.cell(row=1, column=c)
            h.font = Font(name='Calibri', size=11, color="000000", bold=True)
            h.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Coloreo condicional de "Result" (columna J = 10) y wrap en "Error" (columna K = 11)
        from openpyxl.styles import PatternFill
        FILL_GREEN = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")  # verde claro
        FILL_RED   = PatternFill(fill_type="solid", start_color="FFC7CE", end_color="FFC7CE")  # rojo claro

        for r in range(2, max_row + 1):
            result_cell = sheet2.cell(row=r, column=6)  # "Result"
            error_cell  = sheet2.cell(row=r, column=7)  # "Error" (descripción)
            # Asegurar wrap en la descripción de error
            error_cell.alignment = Alignment(wrap_text=True, vertical="top")

            val = str(result_cell.value).strip().lower() if result_cell.value is not None else ""
            if val == "ok":
                result_cell.fill = FILL_GREEN
            elif val == "error":
                result_cell.fill = FILL_RED

        # Ajuste automático de ancho de columnas (con límite)
        for c in range(1, max_col + 1):
            max_len = 0
            for r in range(1, max_row + 1):
                v = sheet2.cell(row=r, column=c).value
                if v is None:
                    continue
                l = len(str(v))
                if l > max_len:
                    max_len = l
            # ancho aproximado: largo + margen; límite para no exagerar
            sheet2.column_dimensions[get_column_letter(c)].width = min(60, max_len + 2)

        # Congelar encabezado
        sheet2.freeze_panes = "A2"

    filepath2 = filepath[:-4] + '.xlsx'
    workbook2.save(filepath2)

    if os.path.exists(filepath):
        os.remove(filepath)
    else:
        print("The file does not exist")

def ImprimirEncabezado():
    nombre_columnas = ['Offering Name' , 'Bono' , 'Codigo Bono', 'Aamount', 'Descripcion', 'Result', 'Error'] 
    Font = "font: name Calibri, color-index black, height 220"
    for i in range(len(nombre_columnas)):
        sheet2.write(0, i, nombre_columnas[i], xlwt.easyxf("pattern: pattern solid, fore_color azul_custom; align: horiz center; "+ Font))
    return 0

def ImprimirEncabezadoPaquetes(j):
    nombre_columnas = ['Offering Name' , 'Client Frendly Name' , 'Modalidad de cobro', 'Action Case Code', 'Marketing Group', 'Result', 'Error'] 
    Font = "font: name Calibri, color-index black, height 220"
    for i in range(len(nombre_columnas)):
        sheet2.write(j, i, nombre_columnas[i], xlwt.easyxf("pattern: pattern solid, fore_color azul_custom; align: horiz center; "+ Font))
    return 0

def separarBonos(value):
    if value is None:
        return []
    if isinstance(value, (list, tuple, set)):
        items = list(value)
    else:
        items = re.split(r"[,\;/\|]+", str(value))  # admite , ; / |
    out = []
    for x in items:
        if x is None:
            continue
        sx = str(x).strip()     # <- convierte ints a str antes de strip
        if sx:
            out.append(sx.upper())
    return out

def ImprimirBase(paquete_bono, bono_nombre, cantidad_bono, unidad_bono,id_bono, j):
    sheet2.write(j, 0, str(paquete_bono)) 
    sheet2.write(j, 1, str(bono_nombre)) 
    sheet2.write(j, 2, str(cantidad_bono)) 
    sheet2.write(j, 3, str(unidad_bono)) 
    sheet2.write(j, 4, str(id_bono))
    
def extraer_valor(campo):
    """
    Recibe un string tipo 'Nombre: valor' y devuelve solo 'valor'
    sin espacios al inicio/fin y sin comillas.
    Si no hay valor (ej: 'Nombre:'), devuelve ''.
    """
    if campo is None:
        return ""

    # Asegurar que es string
    texto = str(campo)

    # Partir por el primer ':'
    partes = texto.split(":", 1)

    if len(partes) < 2:
        # No tiene ':', devolver string limpio
        return texto.strip()

    # Tomar lo que está después de los ':'
    valor = partes[1].strip()

    # Remover comillas simples si las hubiera
    if valor.startswith("'") and valor.endswith("'"):
        valor = valor[1:-1].strip()

    return valor
                                     
    
    
def ValidarDatosPaquetes(paquete_nombre,client_name,modalidad_de_cobro,marketing_group,action_cause_code,paquete):
    paquetes_detalles = _split_comas(paquete[5])
    # print(paquetes_detalles)
    sistema_nombre = paquete[2]
    sistmea_client = extraer_valor(paquetes_detalles[3])
    print(sistmea_client)
    sistema_cobro = extraer_valor(paquetes_detalles[0])
    sistema_cobro = int(sistema_cobro)
    print(sistema_cobro)
    sistema_group = extraer_valor(paquetes_detalles[9])
    print(sistema_group)
    sistema_action_cause_code = extraer_valor(paquetes_detalles[1])
    respuesta = ['OK', '']
    if(str(sistema_nombre).lower() != paquete_nombre):
        respuesta[0] = "Error"
        if(respuesta[1]):
            respuesta[1] += ", Distinto nombre ->"
            respuesta[1]+= str(paquete_nombre)
        else: 
            respuesta[1] = "Distinto nombre->"
            respuesta[1]+= str(paquete_nombre)   
    if(str(sistmea_client).lower() != client_name):
        respuesta[0] = "Error"
        if(respuesta[1]):
            respuesta[1] += ", Distinto client nombre ->"
            respuesta[1]+= str(client_name)
        else: 
            respuesta[1] = "Distinto client nombre->"
            respuesta[1]+= str(client_name)   
    if(sistema_cobro != modalidad_de_cobro):
        respuesta[0] = "Error"
        if(respuesta[1]):
            respuesta[1] += ", Distinta modalidad de cobro ->"
            respuesta[1]+= str(modalidad_de_cobro)
        else: 
            respuesta[1] = "Distinta modalidad de cobro->"
            respuesta[1]+= str(modalidad_de_cobro)   
    if(str(sistema_group).lower() != marketing_group):
        respuesta[0] = "Error"
        if(respuesta[1]):
            respuesta[1] += ", Distinto marketing group ->"
            respuesta[1]+= str(marketing_group)
        else: 
            respuesta[1] = "Distinto marketing group->"
            respuesta[1]+= str(marketing_group)   
    # if(str(sistema_action_cause_code).lower() != action_cause_code):
    #     respuesta[0] = "Error"
    #     if(respuesta[1]):
    #         respuesta[1] += ", Distinto action_cause_code ->"
    #         respuesta[1]+= str(action_cause_code)
    #     else: 
    #         respuesta[1] = "Distinto nombre->"
    #         respuesta[1]+= str(action_cause_code)                                                    

    return respuesta

def buscarBono(bono_nombre, paquete_nombre):
       # Normalizar entradas a minúsculas
    bono_nombre = str(bono_nombre) 
    paquete_nombre = str(paquete_nombre)
    #print(bono_nombre) 
    #print(paquete_nombre)
    for i in range(len(lista_id_bonos_paquetes)): 
        # Normalizar valorpaquetes a minúsculas para la comparación
        Pnombre = str(lista_id_bonos_paquetes[i][1])
        Bnombre = str(lista_id_bonos_paquetes[i][3])
        # Comparar con normalización y verificar si está en desarrollo
        if (bono_nombre == Bnombre and paquete_nombre == Pnombre):
            return lista_id_bonos_paquetes[i]
    
    # Formatear mensaje de error como una cadena
    #mensaje = 'error - No se encontro el Paquete'
    return None

def buscarPaquete(paquete_nombre):
       # Normalizar entradas a minúsculas
    paquete_nombre = str(paquete_nombre).lower() 
    for i in range(len(lista_id_paquetes)): 
        # Normalizar valorpaquetes a minúsculas para la comparación
        Pnombre = str(lista_id_paquetes[i][2]).lower()
        # Comparar con normalización y verificar si está en desarrollo
        if (paquete_nombre == Pnombre):
            return lista_id_paquetes[i]
    
    # Formatear mensaje de error como una cadena
    mensaje = 'error - No se encontro el Paquete'
    return None
            
def _split_comas(value):
    # Acepta str o lista/tupla; devuelve lista limpia de tokens
    if value is None:
        return []
    if isinstance(value, (list, tuple, set)):
        items = list(value)
    else:
        items = str(value).split(",")
    out = []
    for x in items:
        sx = str(x).strip()
        if sx:
            out.append(sx)
    return out           
            
            
def ValidarDatosBonos(bono_nombre, cantidad_bono,unidad_bono,id_bono,vigencia_bono,bono):
    sistema_nombre = bono[3]
    sistema_amount = bono[5]
    sistema_vigencia = bono[4]
    sistema_id_bono = bono[6]
    sistema_descripcion_bono = bono[7]
    sistema_min = bono[8]
    sistema_max = bono[9]
    sistema_default_behavior =bono[10]
    sistema_object_type = bono[12]
    cantidad_bono = float(cantidad_bono)
    respuesta = ['OK', '']
    if(str(sistema_nombre) != bono_nombre):
        respuesta[0] = "Error"
        if(respuesta[1]):
            respuesta[1] += ", Distinto nombre ->"
            respuesta[1]+= str(bono_nombre)
        else: 
            respuesta[1] = "Distinto nombre->"
            respuesta[1]+= str(bono_nombre) 
    if(int(sistema_vigencia) != int(vigencia_bono + 1)):
        respuesta[0] = "Error"
        if(respuesta[1]):
            respuesta[1] += ", Distinta vigencia ->"
            respuesta[1]+= str(vigencia_bono)
        else: 
            respuesta[1] = "Distinto nombre->"
            respuesta[1]+= str(vigencia_bono)
    if(str(sistema_id_bono).lower() != id_bono):
        respuesta[0] = "Error"
        if(respuesta[1]):
            respuesta[1] += ", Distinto id bono ->"
            respuesta[1]+= str(id_bono)
        else: 
            respuesta[1] = "Distinto id bono->"
            respuesta[1]+= str(id_bono) 
    if(str(sistema_descripcion_bono).lower() not in str(bono_nombre).lower()):
        respuesta[0] = "Error"
        if(respuesta[1]):
            respuesta[1] += ", Distinto bonus description ->"
            respuesta[1]+= str(bono_nombre)
        else: 
            respuesta[1] = "Distinto bonus description->"
            respuesta[1]+= str(bono_nombre) 
    if(unidad_bono == 'GB' and sistema_amount != None):
        if(float(str(sistema_amount)) != ((((cantidad_bono *1024)*1024)*1024))):
            print(float(str(sistema_amount)))
            print(((((cantidad_bono *1024)*1024)*1024)))
            respuesta[0] = "Error"
            if(respuesta[1]):
                respuesta[1] += ", Distinto amount ->"
                respuesta[1]+= str(cantidad_bono)
            else: 
                respuesta[1] = "Distinto amount->"
                respuesta[1]+= str(cantidad_bono)
    if(unidad_bono == 'MB' and sistema_amount != None):
        if(float(str(sistema_amount)) != (((cantidad_bono *1024)*1024))):
            respuesta[0] = "Error"
            if(respuesta[1]):
                respuesta[1] += ", Distinto amount ->"
                respuesta[1]+= str(cantidad_bono)
            else: 
                respuesta[1] = "Distinto amount->"
                respuesta[1]+= str(cantidad_bono)
    if(unidad_bono == 'TB' and sistema_amount != None):
        if(float(str(sistema_amount)) != ((((cantidad_bono *1024)*1024)*1024)*1024)):
            respuesta[0] = "Error"
            if(respuesta[1]):
                respuesta[1] += ", Distinto amount ->"
                respuesta[1]+= str(cantidad_bono)
            else: 
                respuesta[1] = "Distinto amount->"
                respuesta[1]+= str(cantidad_bono)                                 
    else:
        if((unidad_bono == 'minutos' or unidad_bono == 'Min'or unidad_bono == 'min') and sistema_amount != None):
            cantidad = (int(cantidad_bono) *60)
            if(int(str(sistema_amount)) != cantidad):
                # print(int(str(sistema_amount)))
                # print(cantidad)
                respuesta[0] = "Error"
                if(respuesta[1]):
                    respuesta[1] += ", Distinto amount ->"
                    respuesta[1]+= str(cantidad_bono)
                else: 
                    respuesta[1] = "Distinto amount->"
                    respuesta[1]+= str(cantidad_bono) 
        else:
            if (unidad_bono == 'sms' and sistema_amount != None):
                if(float(sistema_amount) != float(cantidad_bono)):
                    respuesta[0] = "Error"
                if(respuesta[1]):
                    respuesta[1] += ", Distinto amount ->"
                    respuesta[1]+= str(cantidad_bono)
                else: 
                    respuesta[1] = "Distinto amount->"
                    respuesta[1]+= str(cantidad_bono)
    if(int(sistema_max) != 1 or int(sistema_min) != 1 or sistema_default_behavior != 'Selected'):
        respuesta[0] = "Error"
        if(respuesta[1]):
            respuesta[1] += ", error en la relacion hijo padre->"
            respuesta[1]+= str(sistema_max)
        else: 
            respuesta[1] = "Error en la relacion hijo padre ->"
            respuesta[1]+= str(sistema_max)
    if(sistema_object_type != 'Bolt-On'):
            respuesta[0] = "Error"
            if(respuesta[1]):
                respuesta[1] += ", Distinto object type ->"
                respuesta[1]+= str(sistema_object_type)
            else: 
                respuesta[1] = "Distinto object type->"
                respuesta[1]+= str(sistema_object_type)                         
                                                                                                 
    return respuesta   

def imprimirEquipos(CantidadBonos):
    ''' Esta funcion crear el archivo excel con los equipos'''
    global workbook2, sheet2
    ImprimirEncabezado()
    fila= 2
    salto = 4
    numeroInicio = 2
    Font = "font: name Calibri, color-index black, height 220"
    CantidadBonos = int(CantidadBonos)
    j=1
    for i in range(len(lista_planilla) - 3):
        t=0
        i = i + 3
        k=0
        for t in range(CantidadBonos):
            if(str(lista_planilla[i][19 + (k*5)]) != None):
                bono_nombre = lista_planilla[i][19 + (k*5)]
                #print(bono_nombre)           
                cantidad_bono = (lista_planilla[i][15 + (k*5)])
                #print(cantidad_bono)
                unidad_bono = (lista_planilla[i][16 + (k*5)])
                #print(unidad_bono)
                id_bono = (lista_planilla[i][17 + (k*5)])
                #print(id_bono)
                paquete_nombre = (lista_planilla[i][4])
                #print(paquete_nombre)
                vigencia_bono =  (lista_planilla[i][10])  
                #print(vigencia_bono)
                bono = buscarBono(bono_nombre, paquete_nombre)
                if bono is None:
                    print('No se encontro el bono EXACTO:')
                    print('  Excel paquete =', repr(paquete_nombre))
                    print('  Excel bono    =', repr(bono_nombre))
                    # Mostrar qué bonos tiene el CR para ese mismo paquete
                    print('  Bonos en CR para ese paquete:')
                    for fila in lista_id_bonos_paquetes:
                        if fila[1] == paquete_nombre:
                            print('    -', repr(fila[3]))
                else:
                    ImprimirBase(bono[1], bono_nombre, id_bono,cantidad_bono ,unidad_bono, j)  
                    #ImprimirBaseForPartners(equipo_nombre,equipo[5],revenue_code,prin_color,colores,j)
                    Resultado = ValidarDatosBonos(str(bono_nombre),float(cantidad_bono),str(unidad_bono),str(id_bono).lower(),vigencia_bono,bono)
                    sheet2.write(j, 5, Resultado[0]) #Status
                    sheet2.write(j, 6, Resultado[1]) #Description Error
                    j = j+1
                    k = k+1                    
            else:        
                k = k+1
    p = 0
    ImprimirEncabezadoPaquetes(j)
    j = j + 1
    for p in range(len(lista_planilla) - 3):
        p = p + 3
        paquete_nombre = (lista_planilla[p][4])
        # print(paquete_nombre)
        client_name = lista_planilla[p][5]
        print(client_name)
        modalidad_de_cobro = lista_planilla[p][11]
        print(modalidad_de_cobro)
        marketing_group = lista_planilla[p][((int(cantidad_bonos) + 1 ) *5) + 14]
        print(marketing_group)
        #action_cause_code = lista_planilla[i][((int(cantidad_bonos) + 1 ) *5) + 22]
        paquete = buscarPaquete(paquete_nombre)
        if(paquete == None):
            print('No se encontro el bono', paquete_nombre)
        else:
            ImprimirBase(paquete_nombre, client_name, modalidad_de_cobro, None,marketing_group, j)  
            #ImprimirBaseForPartners(equipo_nombre,equipo[5],revenue_code,prin_color,colores,j)
            Resultado = ValidarDatosPaquetes(str(paquete_nombre).lower(),client_name,int(modalidad_de_cobro),str(marketing_group).lower(),None,paquete)
            sheet2.write(j, 5, Resultado[0]) #Status
            sheet2.write(j, 6, Resultado[1]) #Description Error
            j = j+1                        

    ruta = saveFileDialog()
    workbook2.save(ruta)
    guardarComoXLSX(ruta)


def guardarArchivoText(nombre, objeto):
    '''Funcion guarda archivo en formato texto'''
    archivo = open(nombre,'w') 
    str1 = ',\n'.join(str(e) for e in objeto)
    archivo.write(str1)  
    archivo.close() 
    
    
    
if __name__ == '__main__':
    """Esta la funcion principal""" 
        # Seleccionar opcion para ejecutar script 
    print ('Por favor introduzca el ID del CR:')
    CRID = input()
    print('Introducir cantidad de bonos')
    cantidad_bonos = input()
    lista_id_paquetes = Paquetes(CRID)
    lista_id_bonos = BonosPorParent(CRID)
    lista_id_bonos_paquetes = BonosPorPaqueteDetallado(CRID)
    #lista_id_loyalty = Loyalty(estado)
    now = datetime.datetime.now()
    fecha = now.strftime("%Y-%m-%d %H.%M")
    archivo3 = 'Lista de Paquetes en CR'+fecha+'.txt'
    #guardarArchivoText(archivo2, lista_id_equipos_Online)
    guardarArchivoText(archivo3, lista_id_paquetes)
    archivo2 = 'Lista de Bonos en CR'+fecha+'.txt'
    #guardarArchivoText(archivo2, lista_id_equipos_Online)
    guardarArchivoText(archivo2, lista_id_bonos)
    archivo1 = 'Lista de BonosPaquetes en CR'+fecha+'.txt'
    #guardarArchivoText(archivo2, lista_id_equipos_Online)
    guardarArchivoText(archivo1, lista_id_bonos_paquetes)
    #Solicitar excel de nuevos precios
    ruta = openFileDialog()
    lista_planilla = readCSVPlanes(ruta)    
    imprimirEquipos(cantidad_bonos)    
    print ('Fin')    
       